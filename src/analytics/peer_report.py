import sqlite3
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill


class PeerComparisonReport:

    def __init__(self):

        self.conn = sqlite3.connect(
            "db/nifty100.db"
        )

        self.output_file = (
            "output/peer_comparison.xlsx"
        )

    def load_data(self):

        ratios = pd.read_sql(
            "SELECT * FROM financial_ratios",
            self.conn
        )

        peers = pd.read_sql(
            "SELECT * FROM peer_groups",
            self.conn
        )

        percentiles = pd.read_sql(
            "SELECT * FROM peer_percentiles",
            self.conn
        )

        try:

            companies = pd.read_sql(
                """
                SELECT
                    id,
                    company_name
                FROM companies
                """,
                self.conn
            )

        except Exception:

            companies = pd.read_sql(
                """
                SELECT
                    id,
                    id AS company_name
                FROM companies
                """,
                self.conn
            )

        ratios = (
            ratios
            .sort_values(
                ["company_id", "year"]
            )
            .groupby("company_id")
            .tail(1)
            .reset_index(drop=True)
        )

        df = ratios.merge(
            peers,
            on="company_id",
            how="inner"
        )

        df = df.merge(
            companies,
            left_on="company_id",
            right_on="id",
            how="left"
        )

        return df, percentiles
    
    def build_percentile_table(
        self,
        percentiles
    ):

        pivot = (
            percentiles
            .pivot_table(
                index="company_id",
                columns="metric",
                values="percentile_rank"
            )
            .reset_index()
        )

        pivot.columns.name = None

        return pivot


    def generate_report(self):

        df, percentiles = self.load_data()

        percentile_table = (
            self.build_percentile_table(
                percentiles
            )
        )

        df = df.merge(
            percentile_table,
            on="company_id",
            how="left"
        )

        writer = pd.ExcelWriter(
            self.output_file,
            engine="openpyxl"
        )

        peer_groups = sorted(
            df["peer_group_name"]
            .dropna()
            .unique()
        )

        for group in peer_groups:

            peer_df = (
                df[
                    df["peer_group_name"] == group
                ]
                .copy()
            )

            peer_df.to_excel(
                writer,
                sheet_name=group[:31],
                index=False
            )

        writer.close()
    def format_workbook(self):

        wb = load_workbook(
            self.output_file
        )

        green_fill = PatternFill(
            fill_type="solid",
            start_color="C6EFCE"
        )

        yellow_fill = PatternFill(
            fill_type="solid",
            start_color="FFF2CC"
        )

        red_fill = PatternFill(
            fill_type="solid",
            start_color="F4CCCC"
        )

        gold_fill = PatternFill(
            fill_type="solid",
            start_color="FFD966"
        )

        for sheet in wb.worksheets:

            headers = [
                cell.value
                for cell in sheet[1]
            ]

            percentile_cols = []

            for idx, header in enumerate(headers, start=1):

                if (
                    header is not None
                    and
                    header != "company_id"
                    and
                    (
                        "return_on_equity_pct" in str(header)
                        or
                        "return_on_capital_employed_pct" in str(header)
                        or
                        "net_profit_margin_pct" in str(header)
                        or
                        "debt_to_equity" in str(header)
                        or
                        "free_cash_flow_cr" in str(header)
                        or
                        "pat_cagr_5yr" in str(header)
                        or
                        "revenue_cagr_5yr" in str(header)
                        or
                        "eps_cagr_5yr" in str(header)
                        or
                        "interest_coverage" in str(header)
                        or
                        "asset_turnover" in str(header)
                    )
                ):
                    percentile_cols.append(idx)

            benchmark_column = None

            if "is_benchmark" in headers:
                benchmark_column = (
                    headers.index("is_benchmark") + 1
                )

            for row in range(2, sheet.max_row + 1):

                if benchmark_column:

                    if (
                        sheet.cell(
                            row,
                            benchmark_column
                        ).value == 1
                    ):

                        for col in range(
                            1,
                            sheet.max_column + 1
                        ):
                            sheet.cell(
                                row,
                                col
                            ).fill = gold_fill

                for col in percentile_cols:

                    cell = sheet.cell(row, col)

                    if not isinstance(
                        cell.value,
                        (int, float)
                    ):
                        continue

                    if cell.value >= 75:
                        cell.fill = green_fill

                    elif cell.value <= 25:
                        cell.fill = red_fill

                    else:
                        cell.fill = yellow_fill

            median_row = sheet.max_row + 1

            sheet.cell(
                median_row,
                1
            ).value = "Peer Median"

            for col in range(
                2,
                sheet.max_column + 1
            ):

                values = []

                for row in range(
                    2,
                    median_row
                ):

                    value = sheet.cell(
                        row,
                        col
                    ).value

                    if isinstance(
                        value,
                        (int, float)
                    ):
                        values.append(value)

                if values:

                    sheet.cell(
                        median_row,
                        col
                    ).value = round(
                        pd.Series(values).median(),
                        2
                    )

                wb.save(
            self.output_file
        )

    def run(self):

        print("Loading data...")

        self.generate_report()

        print("Formatting workbook...")

        self.format_workbook()

        print("\n====================================")
        print("peer_comparison.xlsx generated")
        print(f"Location: {self.output_file}")
        print("Sprint 3 Day 20 Complete")
        print("====================================")

        self.conn.close()


def main():

    report = PeerComparisonReport()

    report.run()


if __name__ == "__main__":
    main()