import os

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill

from src.screener.engine import ScreenerEngine

GREEN_FILL = PatternFill(fill_type="solid", start_color="C6EFCE", end_color="C6EFCE")

RED_FILL = PatternFill(fill_type="solid", start_color="FFC7CE", end_color="FFC7CE")


def autosize_columns(ws):

    for column_cells in ws.columns:

        length = max(
            len(str(cell.value)) if cell.value is not None else 0
            for cell in column_cells
        )

        ws.column_dimensions[column_cells[0].column_letter].width = min(length + 3, 35)


def apply_coloring(ws):

    headers = {}

    for cell in ws[1]:
        headers[cell.value] = cell.column

    for row in range(2, ws.max_row + 1):

        # ROE
        if "return_on_equity_pct" in headers:

            cell = ws.cell(row=row, column=headers["return_on_equity_pct"])

            if cell.value is not None:

                if cell.value >= 15:
                    cell.fill = GREEN_FILL
                else:
                    cell.fill = RED_FILL

        # Debt to Equity
        if "debt_to_equity" in headers:

            cell = ws.cell(row=row, column=headers["debt_to_equity"])

            if cell.value is not None:

                if cell.value <= 1:
                    cell.fill = GREEN_FILL
                else:
                    cell.fill = RED_FILL

        # Revenue CAGR
        if "revenue_cagr_5yr" in headers:

            cell = ws.cell(row=row, column=headers["revenue_cagr_5yr"])

            if cell.value is not None:

                if cell.value >= 10:
                    cell.fill = GREEN_FILL
                else:
                    cell.fill = RED_FILL

        # PAT CAGR
        if "pat_cagr_5yr" in headers:

            cell = ws.cell(row=row, column=headers["pat_cagr_5yr"])

            if cell.value is not None:

                if cell.value >= 10:
                    cell.fill = GREEN_FILL
                else:
                    cell.fill = RED_FILL

        # Free Cash Flow
        if "free_cash_flow_cr" in headers:

            cell = ws.cell(row=row, column=headers["free_cash_flow_cr"])

            if cell.value is not None:

                if cell.value >= 0:
                    cell.fill = GREEN_FILL
                else:
                    cell.fill = RED_FILL


def main():

    engine = ScreenerEngine()

    df = engine.load_data()

    presets = [
        "quality_compounder",
        "value_pick",
        "growth_accelerator",
        "dividend_champion",
        "debt_free_blue_chip",
        "turnaround_watch",
    ]

    os.makedirs("output", exist_ok=True)

    output_file = "output/screener_output.xlsx"

    with pd.ExcelWriter(output_file, engine="openpyxl") as writer:

        for preset in presets:

            result = engine.apply_preset(df.copy(), preset)

            result = result.sort_values(by="composite_quality_score", ascending=False)

            result.to_excel(writer, sheet_name=preset[:31], index=False)

    wb = load_workbook(output_file)

    for sheet in wb.sheetnames:

        ws = wb[sheet]

        # Bold headers
        for cell in ws[1]:
            cell.font = Font(bold=True)

        autosize_columns(ws)

        apply_coloring(ws)

    wb.save(output_file)

    engine.conn.close()

    print("screener_output.xlsx generated successfully")
    print("Location:", output_file)


if __name__ == "__main__":
    main()
